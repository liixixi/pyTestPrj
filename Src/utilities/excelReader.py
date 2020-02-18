from openpyxl import load_workbook 

def loadWorkBook(name, columnName):

    workbook = load_workbook(filename=name)
    sheet = workbook.active

    maxRow = sheet.max_row
    titleRowNumber = findTitleRow(sheet, columnName[0])

    dataList = obtainDataList(sheet, maxRow, titleRowNumber, columnName)
    
    return dataList

def findTitleRow(sheet, title):

    for i in range(1, 1000):
        if (sheet['A' + str(i)].value == title):
            return i

def obtainDataList(sheet, maxRow, startRow, columnName):
    
    dataList = []
    for row in range(startRow + 1, maxRow + 1):
        subDataDict = {}
        for name in columnName:
            subDataDict[name] = sheet.cell(row,columnName.index(name) + 1).value
        dataList.append(subDataDict)
    
    return dataList


# testColumnName = ['ID', 'Name', 'Type', 'X', 'Y', 'Height', 'Width', 'Notes']
# datalist = loadWorkBook("Sample.xlsx", testColumnName)
